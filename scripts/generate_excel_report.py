"""
generate_excel_report.py
-------------------------
Builds an Excel workbook from the cleaned trip data with:
  1. Raw_Data        - the cleaned fact-level rows
  2. Payment_Lookup  - a small lookup/dimension table
  3. Summary_Pivot   - a Pandas pivot table (payment_type x hour -> revenue)
  4. VLOOKUP_Demo    - the raw data with REAL Excel VLOOKUP/XLOOKUP formulas
                       (not just pre-computed values) pulling payment_name
                       from Payment_Lookup, so it's demonstrable/editable
                       live in Excel.

Usage:
    python scripts/generate_excel_report.py
Input:
    data/cleaned_tripdata.parquet (produced by data_quality_check.py)
Output:
    data/taxi_analysis_report.xlsx
"""
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

SRC = "data/cleaned_tripdata.parquet"
OUT = "data/taxi_analysis_report.xlsx"

PAYMENT_MAP = {
    1: "Credit card",
    2: "Cash",
    3: "No charge",
    4: "Dispute",
    5: "Unknown",
    6: "Voided trip",
}

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for col in range(1, ncols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def main():
    if not Path(SRC).exists():
        raise FileNotFoundError(f"{SRC} not found — run scripts/data_quality_check.py first.")

    df = pd.read_parquet(SRC).head(1000).copy()  # cap rows for a lightweight demo workbook
    df["hour"] = pd.to_datetime(df["tpep_pickup_datetime"]).dt.hour
    df["row_id"] = range(1, len(df) + 1)

    payment_lookup = pd.DataFrame({
        "payment_type": list(PAYMENT_MAP.keys()),
        "payment_name": list(PAYMENT_MAP.values()),
    })

    pivot = pd.pivot_table(
        df.assign(payment_name=df["payment_type"].map(PAYMENT_MAP)),
        index="hour",
        columns="payment_name",
        values="total_amount",
        aggfunc="sum",
        fill_value=0,
    ).round(2)
    pivot["Grand Total"] = pivot.sum(axis=1).round(2)

    vlookup_cols = ["row_id", "payment_type", "fare_amount", "tip_amount", "total_amount"]
    vlookup_df = df[vlookup_cols].copy()
    vlookup_df["payment_name (VLOOKUP)"] = ""   # filled with real formulas below
    vlookup_df["payment_name (XLOOKUP)"] = ""

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        df.drop(columns=["row_id"]).to_excel(writer, sheet_name="Raw_Data", index=False)
        payment_lookup.to_excel(writer, sheet_name="Payment_Lookup", index=False)
        pivot.to_excel(writer, sheet_name="Summary_Pivot")
        vlookup_df.to_excel(writer, sheet_name="VLOOKUP_Demo", index=False)

        wb = writer.book

        # style headers on every sheet
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            style_header(ws, ws.max_column)

        # --- write REAL Excel formulas into VLOOKUP_Demo ---
        ws = wb["VLOOKUP_Demo"]
        vlookup_col_idx = vlookup_df.columns.get_loc("payment_name (VLOOKUP)") + 1
        xlookup_col_idx = vlookup_df.columns.get_loc("payment_name (XLOOKUP)") + 1
        payment_type_col_idx = vlookup_df.columns.get_loc("payment_type") + 1
        pt_letter = get_column_letter(payment_type_col_idx)

        for r in range(2, len(vlookup_df) + 2):
            ws.cell(row=r, column=vlookup_col_idx).value = (
                f"=VLOOKUP({pt_letter}{r},Payment_Lookup!A:B,2,FALSE)"
            )
            ws.cell(row=r, column=xlookup_col_idx).value = (
                f"=XLOOKUP({pt_letter}{r},Payment_Lookup!A:A,Payment_Lookup!B:B)"
            )

    print(f"Wrote {OUT}")
    print(f"  Raw_Data: {len(df)} rows")
    print(f"  Summary_Pivot: revenue by hour x payment type ({pivot.shape[0]} hours x {pivot.shape[1]} cols)")
    print(f"  VLOOKUP_Demo: {len(vlookup_df)} rows with live VLOOKUP + XLOOKUP formulas "
          f"referencing Payment_Lookup sheet")


if __name__ == "__main__":
    main()
